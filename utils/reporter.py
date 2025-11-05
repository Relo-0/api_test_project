from __future__ import annotations
from pathlib import Path
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class ExcelReporter:
    def __init__(self, out_path: Path):
        self.out_path = out_path

    def write(self, rows: List[Dict[str, Any]]) -> None:
        self.out_path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "API Results"

        headers = [
            "case_name", "method", "url",
            "expected_status", "actual_status",
            "pass", "latency_ms", "error"
        ]
        ws.append(headers)

        for r in rows:
            ws.append([
                r.get("case_name", ""),
                r.get("method", ""),
                r.get("url", ""),
                r.get("expected_status", ""),
                r.get("actual_status", ""),
                "YES" if r.get("pass") else "NO",
                f'{r.get("latency_ms", 0):.1f}',
                r.get("error", "")
            ])

        # 標題樣式
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
            cell.alignment = Alignment(horizontal="center")

        # 自動欄寬
        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max(12, min(max_len + 2, 60))

        wb.save(self.out_path)
